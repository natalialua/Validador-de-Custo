import pandas as pd
import numpy as np
import os
from io import StringIO
from pyxlsb import open_workbook
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font
import re
from datetime import datetime
import win32api
import glob
import win32con

def main(tp_exped, transportadora, file_name_email, file_name_sap, file_name_sap2):

    #################### # 1º DataFrame - Arquivo XLSX #################### 
    folder_path = r"G:"
    file_path = os.path.join(folder_path, file_name_email)
    file_path = os.path.join(folder_path, file_name_email)
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

    if not excel_files:
        raise FileNotFoundError("Nenhum arquivo .xlsx encontrado na pasta.")

    latest_file = max(excel_files, key=os.path.getmtime)

    df = pd.read_excel(latest_file, sheet_name=0, skiprows=14, usecols="B:J", engine="openpyxl")
    df = df.dropna(how='all')

    df.columns = df.columns.str.strip().str.lower()

    df["nota fiscal"] = df["nota fiscal"].astype(str).str.lstrip("0").str.strip()
    df["série"] = df["série"].astype(str).str.lstrip("0").str.strip()
    df["nfe_série"] = df["nota fiscal"] + "-" + df["série"]

    print("Colunas em df:", df.columns)

    #################### # 2º DataFrame - Arquivo TXT #################### 

    file_path = r"G:"
    file_path = os.path.join(folder_path, file_name_sap)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    data = []
    header_line = None

    try:
        with open(file_path, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()
        
        if not lines:
            raise ValueError("O arquivo está vazio!")

        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line):
                header_line = i
                break
        
        if header_line is not None:
            header = [col.strip().lower() for col in lines[header_line].split("|")[1:-1]]

            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line):
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    if row[0] != "*": 
                        data.append(row)

            df_1 = pd.DataFrame(data, columns=header)
            df_1.columns = df_1.columns.str.replace(' ', '').str.strip().str.lower()

            if "notafiscal" in df_1.columns and "série" in df_1.columns:
                    df_1["notafiscal"] = df_1["notafiscal"].astype(str).str.lstrip("0").str.strip()
                    df_1["série"] = df_1["série"].astype(str).str.lstrip("0").str.strip()
                    df_1["nfe_série"] = df_1["notafiscal"] + "-" + df_1["série"]

            # print("Colunas em df_1:", df_1.columns)
        else:
            df_1 = pd.DataFrame()
    except Exception as e:
        print("Erro ao processar o arquivo:", e)
        df_1 = pd.DataFrame()
 
    #################### 3º DataFrame - Arquivo TXT #################### 
    file_path = r"G:"
    file_path = os.path.join(folder_path, file_name_sap2)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"O arquivo {file_path} não foi encontrado.")

    data = []
    header_line = None

    try:
        with open(file_path, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()
        
        if not lines:
            raise ValueError("O arquivo está vazio!")

        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line):
                header_line = i
                break
        
        if header_line is not None:
            header = [col.strip().lower() for col in lines[header_line].split("|")[1:-1]]

            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line):
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    if row[0] != "*":  
                        data.append(row)

            df_2 = pd.DataFrame(data, columns=header)
            df_2.columns = df_2.columns.str.replace(' ', '').str.strip().str.lower()

            if "nfe" in df_2.columns and "série" in df_2.columns:
                df_2["nfe"] = df_2["nfe"].astype(str).str.lstrip("0").str.strip()
                df_2["série"] = df_2["série"].astype(str).str.lstrip("0").str.strip()
                df_2["nfe_série"] = df_2["nfe"] + "-" + df_2["série"]
            else:
                print("⚠️ Aviso: Colunas 'nfe' ou 'série' não encontradas no DataFrame!")

        else:
            df_2 = pd.DataFrame()

    except Exception as e:
        print("❌ Erro ao processar o arquivo:", e)
        df_2 = pd.DataFrame()

    df_2.columns = df_2.columns.str.strip().str.lower()

    # print("✅ Colunas em df_2:", df_2.columns.tolist())

#####################################

    # print("Colunas em df:", df.columns) #depuração
    # print("Colunas em df_2:", df_2.columns)#depuração
    
    df_1 = df_1.rename(columns={
        "u.f.": "u.f",
        "tp.exped.": "Tipo Exped.",
        "descrição": "Destinatário",
        "volume": "Volume m³",
        "série": "Série_df1"  
    })

    df["nfe_série"] = df["nfe_série"].str.strip()
    df_1["nfe_série"] = df_1["nfe_série"].str.strip()

    df_final = df.merge(df_1, on="nfe_série", how="inner")

    df_final = df_final.merge(df_2[['nfe_série', 'senha']], on="nfe_série", how="left", suffixes=('', '_df2'))
    df_final = df_final.merge(df_2[['nfe_série', 'qtdedeestadias']], on="nfe_série", how="left")

    # Renomeando colunas para diferenciar as senhas corretamente
    df_final = df_final.rename(columns={
        "nota fiscal": "Nota Fiscal",
        "série": "Série",
        "cliente destinatário": "Cliente Destinatário",  
        "cidade_x": "Cidade_1", 
        "cidade_y": "Cidade", 
        "u.f_x": "U.F",
        "u.f_y": "UF",
        "senha": "senha_1",  # Senha do df original
        "senha_df2": "senha",  # Senha do df_2
        "tipo do custo": "Tipo do Custo",
        "tipo de veículo": "Tipo de Veículo",
        "qtde veículos": "Qtde Veículos",
        "cod.cliente": "Código Cliente",
        "nro.dtembarque": "DT",
        "nfe_série": "NF + Série",
        "dt.entregacliente": "Data Entrega",
        "qtdedeestadias": "Qtde de Estadias"
    })

###################### NF JÁ AVALIADAS ######################

    def exibir_mensagem_nfs(nf_calculadas):
        """Exibe uma caixa de mensagem com as NFs já calculadas."""
  
        mensagem = "As seguintes Notas Fiscais já foram calculadas e estão no banco de dados:\n" + "\n".join(nf_calculadas)
        win32api.MessageBox(0, mensagem, "NFs Já Calculadas", win32con.MB_ICONINFORMATION)

    # Função para limpar e normalizar as Notas Fiscais
    def limpar_nf(nf):
        """Remove espaços extras e converte para maiúsculas"""
        return str(nf).strip().upper()

    # Caminho para o banco de dados
    source = r"G:"


    def verifica_nf_calculadas(df, conn):
        """
        Verifica se as NFs do DataFrame já foram calculadas e estão armazenadas no banco de dados.
        """
        # Obter as NFs existentes no banco de dados
        query = "SELECT DISTINCT `Nota Fiscal` FROM BI_Estadia"
        nf_existentes = pd.read_sql_query(query, conn)['Nota Fiscal'].dropna().tolist()
        
        # Normalizar as NFs do banco
        nf_existentes = [limpar_nf(nf) for nf in nf_existentes]
        
        # Normalizar as NFs no DataFrame
        df['Nota Fiscal'] = df['Nota Fiscal'].apply(limpar_nf)
        
        # Verificar se há NFs duplicadas
        nf_calculadas = df[df['Nota Fiscal'].isin(nf_existentes)]
        
        if not nf_calculadas.empty:
            # Exibir a mensagem usando a função de toast
            exibir_mensagem_nfs(nf_calculadas['Nota Fiscal'].unique())
            # Remover as NFs já calculadas do DataFrame
            df = df[~df['Nota Fiscal'].isin(nf_existentes)]
        
        return df, nf_calculadas['Nota Fiscal'].unique()

    # Conectar ao banco de dados
    conn = sqlite3.connect(source)

    # Verificar NFs já calculadas antes de gerar o arquivo
    df_final, nfs_avaliadas = verifica_nf_calculadas(df_final, conn)

    if df_final.empty:
        print("Todas as Notas Fiscais já foram calculadas.")
    else:
        print("Continuando o processo de geração de arquivo...")

    ####CALCULOS#####

    df_final['Data Entrega'] = pd.to_datetime(df_final['Data Entrega'], errors='coerce', dayfirst=True)
    df_final['Data Entrega'] = df_final['Data Entrega'].fillna(pd.Timestamp('1900-01-01'))
    df_final['Data Entrega'] = pd.to_datetime(df_final['Data Entrega'], errors='coerce', dayfirst=True)
    df_final['dt.numero'] = (df_final['Data Entrega'] - pd.Timestamp('1900-01-01')).dt.days + 2
    df_final["Chave de Entrega"] = df_final["dt.numero"].astype(str) + df_final["Código Cliente"].astype(str)
    df_final['Data Entrega'] = df_final['Data Entrega'].dt.strftime('%d/%m/%Y')
    df_final = df_final.drop(columns=['dt.numero'])

    #REQ
    def get_next_filename(base_name, extension, folder_path):
        i = 0
        while os.path.exists(os.path.join(folder_path, f"{base_name + i}.{extension}")):
            i += 1
        return base_name + i

    base_name = 555
    extension = "txt"
    folder_path = r"G:"

    next_filename = get_next_filename(base_name, extension, folder_path)
    df_final['Requisição'] = next_filename

    df_final["Qtd veiculos aprovados"] = (df_final["Chave de Entrega"] != df_final["Chave de Entrega"].shift(-1)).astype(int)
    df_final["Volume m³"] = df_final["Volume m³"].str.replace(',', '.').astype(float)
    df_final["Soma Vol m³"] = df_final.groupby("Chave de Entrega")["Volume m³"].cumsum()
    df_final["Soma Vol m³"] = df_final["Soma Vol m³"] * (df_final["Qtd veiculos aprovados"] > 0)
    df_final["Soma Vol m³"] = df_final.groupby("Chave de Entrega")["Soma Vol m³"].transform("max")
    df_final.loc[df_final["Qtd veiculos aprovados"] == 0, "Soma Vol m³"] = 0
    df_final['Data Requisição'] = datetime.today().strftime('%d/%m/%Y')

    # Conectar ao banco SQLite
    conn = sqlite3.connect(r"G:"
)  
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info('valor_dedicado');")
    colunas = cursor.fetchall()
    cursor.execute("SELECT * FROM valor_dedicado")
    valor_dedicado_data = cursor.fetchall()
    valor_dedicado_df = pd.DataFrame(valor_dedicado_data, columns=["Tipo veiculo", "min", "max", "valor"])
    valor_dedicado_df["min"] = pd.to_numeric(valor_dedicado_df["min"], errors="coerce")
    valor_dedicado_df["max"] = pd.to_numeric(valor_dedicado_df["max"], errors="coerce")

    def validar_tipo_veiculo(volume):
        volume = pd.to_numeric(str(volume).replace(',', '.'), errors='coerce')
        
        if pd.isna(volume): 
            return None

        print(f"Validando volume: {volume}")
        
        if volume < 1:
            print(f"Volume {volume} é menor que 1, assumindo 'FIORINO'")
            return "FIORINO"
        if volume > 55:
            print(f"Volume {volume} é maior que 55, assumindo 'CARRETA'")
            return "CARRETA"

        for _, row in valor_dedicado_df.iterrows():
            if pd.notna(row["min"]) and pd.notna(row["max"]):  
                if row["min"] <= volume <= row["max"]:
                    print(f"Volume {volume} está entre {row['min']} e {row['max']}, Tipo veiculo: {row['Tipo veiculo']}")
                    return row["Tipo veiculo"]

        return None  


    # df_final["Tipo de veículo validado"] = df_final["Soma Vol m³"].apply(validar_tipo_veiculo)

    df_final["Tipo de veículo validado"] = ""

    for chave, grupo in df_final.groupby("Chave de Entrega"):
        soma_vol = grupo["Soma Vol m³"].sum()
        tipo_veiculo = validar_tipo_veiculo(soma_vol)

        idx_ultimo = grupo.index[-1]

        df_final.at[idx_ultimo, "Tipo de veículo validado"] = tipo_veiculo

    # print("Dados do df_final após validação:")
    # print(df_final[["Soma Vol m³", "Tipo de veículo validado"]])

    transp_df = pd.read_sql("SELECT * FROM Transp", conn)
    # print(transp_df["De"].value_counts())
    df_final = df_final.rename(columns={"dsctrans.entr.cliente": "De"})
    df_final = df_final.merge(transp_df[['De', 'PARA']], on='De', how='left')
    df_final = df_final.rename(columns={"PARA": "Transportador"})

    # df_final = df_final[(df_final['Transportador'] == transportadora)]
    # df_final = df_final[(df_final['Tipo Exped'] == tp_exped)]
    
    df_final = df_final[ (df_final['Transportador'] == transportadora)]
    
    ordem_final = [
        "Data Requisição", "Nota Fiscal", "Série", "Cliente Destinatário", "Cidade_1", "U.F",
        "Tipo do Custo", "Tipo de Veículo", "Qtde Veículos", "senha_1",  "Requisição", 
        "Código Cliente", "Destinatário", "Cidade", "UF", "DT", "NF + Série", 
        "Transportador", "Data Entrega", "Tipo Exped.", "Chave de Entrega", "Volume m³", 
        "Qtd veiculos aprovados", "Soma Vol m³", "senha", "Qtde de Estadias", 
        "Tipo de veículo validado", "Valor Validado ", "Valor Total"
    ]

    df_final = df_final[[col for col in ordem_final if col in df_final.columns]]

    # print(df_final.columns)#depuração

    # Obtendo os dados da tabela 'tabelas_fracionado'
    tabelas_fracionado_df = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)

    # 🔹 Mapeamento de veículos
    mapeamento_veiculos = {
        "FIORINO": "ESTADIA FIORINO",
        "VAN": "ESTADIA VAN",
        "VUC": "ESTADIA VUC",
        "3/4": "ESTADIA 3/4",
        "TOCO": "ESTADIA TOCO",
        "TRUCK": "ESTADIA TRK",
        "CARRETA": "ESTADIA CAR"
    }

    df_final["Valor Validado"] = None

    for chave in df_final["Chave de Entrega"].unique():
        subset = df_final[df_final["Chave de Entrega"] == chave]
        
        idx = subset.index[subset["Qtd veiculos aprovados"] > 0].min()
        
        df_final.loc[subset.index, ["Valor Validado ", "Valor Total"]] = None
        
        if not pd.isna(idx):
            transportadora = df_final.loc[idx, "Transportador"]
            tipo_veiculo = df_final.loc[idx, "Tipo de veículo validado"]
            coluna_veiculo = mapeamento_veiculos.get(tipo_veiculo)
            
            if transportadora in tabelas_fracionado_df["TRANSPORTADORA"].values and coluna_veiculo:
                valor = tabelas_fracionado_df.loc[
                    tabelas_fracionado_df["TRANSPORTADORA"] == transportadora, coluna_veiculo
                ].values[0]
                df_final.at[idx, "Valor Validado"] = valor

    df_final["Qtde de Estadias"] = pd.to_numeric(df_final["Qtde de Estadias"], errors="coerce")
    df_final["Valor Validado "] = pd.to_numeric(df_final["Valor Validado"], errors="coerce")

    df_final["Qtde de Estadias"] = pd.to_numeric(df_final["Qtde de Estadias"], errors="coerce")
    df_final["Valor Validado "] = pd.to_numeric(df_final["Valor Validado"], errors="coerce")

    df_final["Valor Total"] = df_final["Qtde de Estadias"] * df_final["Valor Validado"]

    soma_valor = df_final["Valor Total"].dropna().sum()
    
    def limpa_e_numeric(serie):
        s = serie.astype(str).str.strip()
        s = s.str.replace(r"[^0-9,\.-]", "", regex=True)
        s = s.str.replace(",", ".", regex=False)
        return pd.to_numeric(s, errors="coerce").fillna(0)
    
    colunas_valores = [
        "Valor Total"  
    ]

    # 🔹 Converte as colunas numéricas
    for col in colunas_valores:
        if col in df_final.columns:
            df_final[col] = limpa_e_numeric(df_final[col])

    # 🔹 Calcula totais das colunas válidas
    totais = {col: df_final[col].sum() for col in colunas_valores if col in df_final.columns}

    df_nfs_avaliadas = pd.DataFrame(nfs_avaliadas, columns=['Nota Fiscal'])

    # Caminho do arquivo Excel
    output_xlsx = r"G:"
    output_txt = r"G:"

    with pd.ExcelWriter(output_xlsx, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, sheet_name="Relatório", startrow=1, index=False)
        df_nfs_avaliadas.to_excel(writer, sheet_name="nf avaliadas", index=False)

    # 🔹 **Aplicar formatação 
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Relatório"]  

    # estilos
    fill_header = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    font_branca = Font(color="FFFFFF")
    font_yellow = Font(bold=True, color="FFFF00")
    for col in colunas_valores:
        idx = df_final.columns.get_loc(col) + 1  
        cell = ws.cell(row=1, column=idx)
        texto = f"R$ {totais[col]:,.2f}"
        # adapta para PT‑BR
        texto = texto.replace(",", "v").replace(".", ",").replace("v", ".")
        cell.value = texto
        cell.fill  = fill_header
        cell.font  = font_yellow
    for cell in ws[2]:
        cell.fill = fill_header
        cell.font = font_branca
       

    wb.save(output_xlsx)
    print(f"Arquivo Excel salvo em: {output_xlsx}")

    # 🔹 **Salvar como TXT**
    df_final.to_csv(output_txt, sep="|", index=False, encoding="utf-8")
    print(f"Arquivo TXT salvo em: {output_txt}")

if __name__ == "__main__":
    pass


