import pandas as pd
import numpy as np
import os
from io import StringIO
from pyxlsb import open_workbook
import re
import sqlite3
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from win10toast import ToastNotifier
import win32api
import win32con

def main(tp_exped, transportadora, file_name):

    folder_path = r'G:'
    file_path = os.path.join(folder_path, file_name) 
    data = []

    try:
        with open(file_path, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()

        source = os.path.dirname(file_path)  
        arquivos = os.listdir(source)

    except Exception as e:
        print("Erro ao processar o arquivo:", e)

########### Leitura ZSD173 ##################
    # Caminhos do arquivo .txt e do arquivo Excel
    caminho_txt = r"G:"
    with open(caminho_txt, 'r', encoding='latin1') as file:
        linhas = file.readlines()
    linhas_dados = [
        linha for linha in linhas if linha.startswith('|') and not linha.strip().startswith('---')
    ]
    cabecalho = linhas_dados.pop(0).strip('|').split('|')
    colunas = [col.strip() for col in cabecalho]
    dados = [linha.strip('|').split('|') for linha in linhas_dados]

    df_173 = pd.DataFrame(dados, columns=colunas)

    df_173 = df_173.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # Criar a nova coluna "Chave" como a junção de "Nr.Nota Fi" e "Series."
    if "Nr.Nota Fi" in df_173.columns and "Series." in df_173.columns:
        df_173.insert(0, "Chave", df_173["Nr.Nota Fi"] + "-" + df_173["Series."])

    df_173 = df_173.drop_duplicates().reset_index(drop=True)

    ###################### LEITURA ZSD223 ######################

    header_line = None
    data = []
    try:
        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line): 
                header_line = i
                break

        if header_line is not None:
            header = [col.strip() for col in lines[header_line].split("|")[1:-1]]

            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line): 
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    data.append(row)

            df = pd.DataFrame(data, columns=header)
            df.columns = df.columns.str.replace(' ', '')
            df.columns = df.columns.str.strip()
            # print(df.head())  
        else:
            print("Cabeçalho não encontrado no arquivo!")
    except Exception as e:
        print("Erro ao processar o arquivo:", e)
    df.columns = df.columns.str.strip()

    ## CONSULTAS SQL
        
    conn = sqlite3.connect(r"")  
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info('TABELAS_FRACIONADO');")
    colunas = cursor.fetchall()

    cursor.execute("PRAGMA table_info('TABELAS_LOTACAO');")
    colunas = cursor.fetchall()

    cursor.execute("PRAGMA table_info(Book);")
    ##print([coluna[1] for coluna in cursor.fetchall()])

    cursor.execute("PRAGMA table_info(transp);")
    ##print([coluna[1] for coluna in cursor.fetchall()])

    cursor.execute("PRAGMA table_info(Base_Produtos);")
    ##print([coluna[1] for coluna in cursor.fetchall()])

    cursor.execute("PRAGMA table_info(Base_Clientes);")
    ##print([coluna[1] for coluna in cursor.fetchall()])

    cursor.execute("PRAGMA table_info(Info_Gerais);")

    ##SELECT 
    book_df = pd.read_sql("SELECT * FROM Book", conn)
    transp_df = pd.read_sql("SELECT * FROM Transp", conn)
    prod_df = pd.read_sql("SELECT * FROM Base_Produtos", conn)
    cliente_df = pd.read_sql("SELECT* FROM Base_Clientes", conn)
    info_df = pd.read_sql("SELECT* FROM Info_Gerais", conn)
    fraci_df = pd.read_sql("SELECT* FROM TABELAS_FRACIONADO", conn)
    lota_df = pd.read_sql("SELECT* FROM TABELAS_LOTACAO", conn)

   ###################### PROCESSAMENTO_ZSD223 ######################

    ##NFE + Série
    df['NFE + Série'] = df['NotaFiscal'].astype(str) + '-' + df['Série'].astype(str)
    df.columns = df.columns.str.strip()

    ##TP
    df_merged = df.merge(transp_df[['De', 'PARA']], 
                        left_on=['DscTrans.Entr.Cliente'], 
                        right_on=['De'], 
                        how='left')
    df_merged['TP'] = df_merged['PARA'].fillna('-')

    # Remova as colunas 'De' e 'PARA' do DataFrame resultante
    df_merged = df_merged.drop(columns=['De', 'PARA'])

    #Conversões
    try:
    
        df_merged['Quant.Fornecida'] = df_merged['Quant.Fornecida'].str.replace('.', '').str.replace(',', '.').replace('', np.nan).astype(float)

    except:

        df_merged['Quant.Fornecida'] = df_merged['Quant.Fornecida'].replace('.', '').replace(',', '.').replace('', np.nan).astype(float)

    ##Tipo de Paletização
    def get_tipo_paletizacao(cod_cliente):
        try:
            if book_df.loc[book_df['CÓDIGO'] == cod_cliente, 'PALETIZAÇÃO DIFERENCIADA?'].values[0] == 'PROCEDE':
                return 'PADRÃO CLIENTE'
            else:
                return 'PADRÃO'
        except IndexError:
            return 'PADRÃO'

    df_merged['Tipo de Paletização'] = df_merged['Cod.Cliente'].apply(get_tipo_paletizacao)

    ##UMB
    df_merged['UMB'] = df_merged['Material'].map(prod_df.drop_duplicates(subset='Material').set_index('Material')['UMB']).fillna('-')

    ##Unid. Caixa
    df_merged['Unid. Caixa'] = df_merged['Material'].map(prod_df.drop_duplicates(subset='Material').set_index('Material')['Unidades CX']).fillna('-')
    df_merged['Unid. Caixa'] = pd.to_numeric(df_merged['Unid. Caixa'].replace('', 'NaN'), errors='coerce')

    ##Base 2 form. Qdt conv cx
    def calcular_base_2(row):
        try:
            if row['UMB'] == 'UN':  
                return row['Quant.Fornecida'] / row['Unid. Caixa']
            else:  
                return row['Quant.Fornecida']
        except (KeyError, ZeroDivisionError, TypeError):
            return '-'  
    df_merged['Base 2 form. Qdt conv cx'] = df_merged.apply(calcular_base_2, axis=1)

    ##Base 1 form. Qdt conv cx
    def calcular_base_1(row):
        try:
            if row['UMB'] == 'CX' and row['Base 2 form. Qdt conv cx'] == row['Quant.Fornecida']:
                material = row['Material']
                valor = prod_df.loc[prod_df['Material'] == material, 'Unidades (UN)']
                if not valor.empty:
                    return valor.values[0] 
                else:
                    return 0  
            else:
                return 0  
        except (KeyError, TypeError):
            return 0  
    df_merged['Base 1 form. Qdt conv cx'] = df_merged.apply(calcular_base_1, axis=1)

    ##Camadas 
    df_merged['Material'] = df_merged['Material'].astype(str)
    prod_df['Material'] = prod_df['Material'].astype(str)
    df_merged = df_merged.merge(prod_df[['Material', 'Camadas']], 
                        left_on=['Material'], 
                        right_on=['Material'], 
                        how='left')
    def calcular_camadas(row):
        try:
            if row['UMB'] == 'UN':
                return row['Camadas'] / row['Unid. Caixa']
            else:
                return row['Camadas']
        except:
            return '-'

    df_merged['Camadas'] = df_merged.apply(calcular_camadas, axis=1)

    ##Lastro
    def calcular_lastro(row):
        try:
            if row['UMB'] == 'UN':
                return float(prod_df.loc[prod_df['Material'] == row['Material'], 'Lastro'].values[0]) / float(row['Unid. Caixa'])
            else:
                return float(prod_df.loc[prod_df['Material'] == row['Material'], 'Lastro'].values[0])
        except (IndexError, ValueError):
            return '-'

    df_merged['Lastro'] = df_merged.apply(calcular_lastro, axis=1)
    df_merged['Lastro'] = pd.to_numeric(df_merged['Lastro'], errors='coerce')

    ##Qtd Caixas por palete
    df_merged['Qtd Caixas por palete'] = (df_merged['Lastro'] * df_merged['Camadas']).fillna('-')

    ##Qtd. Paletes

    df_merged['Unid. Caixa'] = pd.to_numeric(df_merged['Unid. Caixa'].replace('', np.nan), errors='coerce')
    df_merged['Qtd Caixas por palete'] = pd.to_numeric(df_merged['Qtd Caixas por palete'].replace('', np.nan), errors='coerce')

    def calcular_qtd_paletes(row):
        try:
            quant_fornecida = row['Quant.Fornecida']
            unid_caixa = row['Unid. Caixa']
            qtd_caixas_por_palete = row['Qtd Caixas por palete']
            
            if pd.isna(quant_fornecida) or pd.isna(qtd_caixas_por_palete):
                return '-'
            if row['UMB'] == 'UN':
                if pd.isna(unid_caixa):
                    return '-'
                return quant_fornecida / (unid_caixa * qtd_caixas_por_palete)
            else:
                return quant_fornecida / qtd_caixas_por_palete
        except (ZeroDivisionError, TypeError):
            return '-'
    df_merged['Qtd. Paletes'] = df_merged.apply(calcular_qtd_paletes, axis=1)

    ##Altura CX
    df_merged['Altura CX'] =df_merged['Material'].map(prod_df.drop_duplicates(subset='Material').set_index('Material')['Altura  CX']).fillna('-')
    df_merged['Altura CX'] = pd.to_numeric(df_merged['Altura CX'], errors='coerce')

    ##Altura Palete
    df_merged = df_merged.merge(prod_df[['Material', 'Altura Total']], on='Material', how='left')
    df_merged = df_merged.rename(columns={'Altura Total': 'Altura Palete'})

    ##Qtd Sku Palete CLI
    df_merged['Qtd Sku Palete CLI'] = df_merged['Cod.Cliente'].map(book_df.drop_duplicates(subset='CÓDIGO').set_index('CÓDIGO')['QTD SKU ']).fillna('-')
    df_merged['Qtd Sku Palete CLI'] = pd.to_numeric(df_merged['Qtd Sku Palete CLI'], errors='coerce')

    #Altura Máx CLI
    df_merged['Altura Palete'] = pd.to_numeric(df_merged['Altura Palete'], errors='coerce')
    df_merged['Qtd Sku Palete CLI'] = pd.to_numeric(df_merged['Qtd Sku Palete CLI'], errors='coerce')
    cliente_df['Altura Máxima Carga/Palete (mt)'] = pd.to_numeric(cliente_df['Altura Máxima Carga/Palete (mt)'], errors='coerce')

    df_merged['Cod.Cliente'] = df_merged['Cod.Cliente'].astype(str)
    cliente_df['Cliente'] = cliente_df['Cliente'].astype(str)

    df_merged = df_merged.merge(cliente_df[['Cliente', 'Altura Máxima Carga/Palete (mt)']],left_on='Cod.Cliente',right_on='Cliente',how='left')
    df_merged = df_merged.rename(columns={'Altura Máxima Carga/Palete (mt)': 'Altura Máx CLI'})
    df_merged['Altura Máx CLI'] = df_merged['Altura Máx CLI'].fillna(df_merged['Altura Palete'])

    ##Qtd Sku Palete CLI
    df_merged['Qtd Sku Palete CLI'] = df_merged['Cod.Cliente'].map(book_df.drop_duplicates(subset='CÓDIGO').set_index('CÓDIGO')['QTD SKU ']).fillna('-')
    df_merged['Qtd Sku Palete CLI'] = pd.to_numeric(df_merged['Qtd Sku Palete CLI'], errors='coerce')

    ##CLI Aceita Pal Misto?
    df_merged['CLI Aceita Pal Misto?'] =df_merged['Qtd Sku Palete CLI'].apply(lambda x: 'SIM' if x > 0 else 'NÃO')

    # 'Camada CLI'
    df_merged["Camada CLI"] = np.where(df_merged["Altura Máx CLI"] == 0, df_merged["Camadas"],  
        np.minimum(
            np.floor(df_merged["Altura Máx CLI"] / df_merged["Altura CX"]), 
            df_merged["Camadas"]
        )
    )

    df_merged["Camada CLI"] = df_merged["Camada CLI"].fillna("-")
    df_merged['CLI Aceita Pal Misto?'] = df_merged['Qtd Sku Palete CLI'].apply(lambda x: 'SIM' if x > 0 else 'NÃO')

    ##Qtd Caixas / Palete CLI
    df_merged['Qtd Caixas / Palete CLI'] = (df_merged['Lastro'] * df_merged['Camada CLI']).fillna('-')

    ##Altura Palete CLI
    df_merged['Altura Palete CLI'] = (df_merged['Altura CX'] * df_merged['Camada CLI']).fillna('-')

    ##UV
    df_merged['NFE + Série'] = df_merged['NFE + Série'].str.lstrip('0')

    df_merged = df_merged.merge(df_173[['Chave', 'Material', 'UV']], 
                                left_on=['NFE + Série', 'Material'], 
                                right_on=['Chave', 'Material'], 
                                how='left')

    df_merged.drop(columns=['Chave'], inplace=True)

    ## Qtd Fornecida (Convert. Cxs)
    df_merged['Qtd Caixas por palete'] = pd.to_numeric(df_merged['Qtd Caixas por palete'].replace('', np.nan), errors='coerce')

    # Função para calcular 'Qtd Fornecida (Convert. Cxs)'
    def calcular_qtd_fornecida(row):
        try:
            if row['UV'] == 'CX': 
                return row['Quant.Fornecida']
            elif row['Base 1 form. Qdt conv cx'] > 0: 
                return row['Base 1 form. Qdt conv cx']
            else: 
                return row['Base 2 form. Qdt conv cx']
        except (KeyError, TypeError):
            return '-' 
    df_merged['Qtd Fornecida (Convert. Cxs)'] = df_merged.apply(calcular_qtd_fornecida, axis=1)

    ##Sequencia NF
    df_merged['Sequencia NF'] = df_merged['Série'].apply(lambda x: 0 if x == 6 else x)

    ##Cli+codcriado
    df_merged['Cli+codcriado'] = df_merged['Cod.Cliente'].astype(str) + df_merged['Série'].apply(lambda x: '0' if x == 6 else '1')

    ## PALETES VALIDADOS
    def calcular_nova_coluna(row):
        try:
            if row['Tipo de Paletização'] == 'PADRÃO ':
                return 0
            elif row['CLI Aceita Pal Misto?'] == 'NÃO':
                return math.ceil(row['Qtd Fornecida (Convert. Cxs)'] / row['Qtd Caixas / Palete CLI'])
            else:
                return math.ceil((row['Qtd Fornecida (Convert. Cxs)'] / row['Qtd Caixas / Palete CLI']) * 10) / 10
        except (KeyError, TypeError, ZeroDivisionError):
            return '-'
    df_merged['PALETES VALIDADOS'] = df_merged.apply(calcular_nova_coluna, axis=1)

# REQ
    # Função para encontrar o próximo nome de arquivo disponível
    def get_next_filename(base_name, extension, folder_path):
        i = 0
        while os.path.exists(os.path.join(folder_path, f"{base_name + i}.{extension}")):
            i += 1
        return base_name + i

    # Base name, extensão e caminho da pasta
    base_name = 668
    extension = "txt"
    folder_path = r"G:"

    # Obter o próximo nome de arquivo disponível
    next_filename = get_next_filename(base_name, extension, folder_path)

    # Atualizar o valor de REQ
    df_merged['REQ'] = next_filename

    ##Chave
    df_merged['Dt.EntregaCliente'] = pd.to_datetime(df_merged['Dt.EntregaCliente'], errors='coerce', dayfirst=True)


    df_merged['Dt.EntregaCliente'] = df_merged['Dt.EntregaCliente'].fillna(pd.Timestamp('1900-01-01'))

    df_merged['Dt.Numero'] = ((df_merged['Dt.EntregaCliente'] - pd.Timestamp('1900-01-01')).dt.days + 2).astype(int)

    df_merged['Chave'] = df_merged['Cod.Cliente'].astype(str) + \
                        df_merged['Dt.Numero'].astype(str) + \
                        df_merged['REQ'].astype(str)

    df_merged = df_merged.drop(columns=['Dt.Numero'])

    num_linhas = df_merged.shape[0]
    print(f'O número de linhas é: {num_linhas}')

    ##Filtros 

    df_merged['Tp.exped.'] = df_merged['Tp.exped.'].str.strip()
    df_merged['TP'] = df_merged['TP'].str.strip()

    df_merged = df_merged[df_merged['Tp.exped.'] != '']

    df_merged = df_merged[(df_merged['Tp.exped.'] == tp_exped) & (df_merged['TP'] == transportadora)]

###################### Colunas ######################
    
    df_merged.columns = ['Localorg.transporte', 'Cod.Cliente', 'Descrição', 'Cidade', 'U.F.',
                        'NotaFiscal', 'Série', 'Material', 'Descrição2', 'Quant.Fornecida',
                        'ValorNF', 'Volume', 'VolumeNF', 'Peso', 'Nro.DTEmbarque',
                        'DscTrans.Entr.Cliente', 'Dt.EntregaCliente', 'Tp.exped.',
                        'Proc.especial', 'NFE + Série', 'TP', 'Tipo de Paletização', 'UMB',
                        'Unid. Caixa', 'Base 2 form. Qdt conv cx', 'Base 1 form. Qdt conv cx',
                        'Camadas', 'Lastro', 'Qtd Caixas por palete', 'Qtd. Paletes',
                        'Altura CX', 'Altura Palete', 'Qtd Sku Palete CLI', 'Cliente',
                        'Altura Máx CLI', 'CLI Aceita Pal Misto?', 'Camada CLI',
                        'Qtd Caixas / Palete CLI', 'Altura Palete CLI', 'UV',
                        'Qtd Fornecida (Convert. Cxs)', 'Sequencia NF', 'Cli+codcriado',
                        'PALETES VALIDADOS', 'REQ', 'Chave']
   
    df_plt = df_merged[['REQ', 'Tipo de Paletização', 'Cod.Cliente', 'Descrição', 'TP', 'Dt.EntregaCliente', 'NFE + Série', 'PALETES VALIDADOS', 'Tp.exped.']].copy()
    
    df_plt.rename(columns={
        'REQ': 'Solicitação',
        'Tipo de Paletização': 'Perfil Cliente',
        'Cod.Cliente': 'Cod.Cliente',
        'Descrição': 'Descrição',
        'TP': 'Transportador',
        'Dt.EntregaCliente': 'Dt/ Entrega Cliente',
        'NFE + Série': 'Nota fiscal',
        'PALETES VALIDADOS': 'Paletes validados',
        'Tp.exped.':  'Tp.exped.'
    }, inplace=True)


###################### NF JÁ AVALIADAS ###############
    # Exibir mensagem após a execução bem-sucedida
    def exibir_mensagem_nfs(nf_calculadas):
        """Exibe uma caixa de mensagem com as NFs já calculadas."""
        mensagem = "As seguintes Notas Fiscais já foram calculadas e estão no banco de dados:\n" + "\n".join(nf_calculadas)
        win32api.MessageBox(0, mensagem, "NFs Já Calculadas", win32con.MB_ICONINFORMATION)

    # Função para limpar e normalizar as Notas Fiscais
    def limpar_nf(nf):
        """Remove espaços extras e converte para maiúsculas"""
        return str(nf).strip().upper()

    # Caminho para o banco de dados
    source = r""

    def verifica_nf_calculadas(df, conn):
        """
        Verifica se as NFs do DataFrame já foram calculadas e estão armazenadas no banco de dados.
        """
        # Obter as NFs existentes no banco de dados
        query = "SELECT DISTINCT `Nota fiscal` FROM BI_Paletizacao"
        nf_existentes = pd.read_sql_query(query, conn)['Nota fiscal'].dropna().tolist()
        
        # Normalizar as NFs do banco
        nf_existentes = [limpar_nf(nf) for nf in nf_existentes]
        
        # Normalizar as NFs no DataFrame
        df['Nota fiscal'] = df['Nota fiscal'].apply(limpar_nf)
        
        # Verificar se há NFs duplicadas
        nf_calculadas = df[df['Nota fiscal'].isin(nf_existentes)]
        
        if not nf_calculadas.empty:
            # Exibir a mensagem usando a função de toast
            exibir_mensagem_nfs(nf_calculadas['Nota fiscal'].unique())
            # Remover as NFs já calculadas do DataFrame
            df = df[~df['Nota fiscal'].isin(nf_existentes)]
        
        return df, nf_calculadas['Nota fiscal'].unique()

    # Conectar ao banco de dados
    conn = sqlite3.connect(source)

    # Verificar NFs já calculadas antes de gerar o arquivo
    df_plt, nfs_avaliadas = verifica_nf_calculadas(df_plt, conn)

    if df_plt.empty:
        print("Todas as Notas Fiscais já foram calculadas.")
    else:
        print("Continuando o processo de geração de arquivo...")


#################### CALCULOS ####################

    df_plt['Perfil Cliente'] = df_plt['Perfil Cliente'].apply(lambda x: 'Procede' if x == 'PADRÃO CLIENTE' else 'Não procede')

    df_plt['Dt/ Entrega Cliente'] = pd.to_datetime(df_plt['Dt/ Entrega Cliente'], errors='coerce').dt.strftime('%d/%m/%Y')
    df_plt['Data Validação'] = datetime.today().strftime('%d/%m/%Y')

    df_plt['Paletes validados'] = pd.to_numeric(df_plt['Paletes validados'], errors='coerce')

    # Escolhe o DataFrame correto com base em tp_exped
    if tp_exped == 'Z2':
        df_base = fraci_df
    elif tp_exped in ['Z1', 'E1']:
        df_base = lota_df
    else:
        raise ValueError("Tipo de expedição inválido. Esperado 'Z1', 'Z2' ou 'E1'")

    valor_paletizacao = df_base.loc[df_base['TRANSPORTADORA'] == transportadora, 'TX PALETIZAÇÃO'].values[0]
  
    df_plt['Valor Aprovado'] = df_plt['Paletes validados'] * valor_paletizacao

    df_plt = df_plt.groupby(
        ['Solicitação', 'Perfil Cliente', 'Data Validação', 'Cod.Cliente', 'Descrição', 'Transportador', 'Dt/ Entrega Cliente', 'Nota fiscal', 'Tp.exped.'],
        as_index=False
    ).agg({'Paletes validados': 'sum', 'Valor Aprovado': 'sum'})

    df_total = df_plt.groupby(
        ['Solicitação', 'Perfil Cliente', 'Data Validação', 'Cod.Cliente', 'Descrição', 'Transportador', 'Dt/ Entrega Cliente', 'Tp.exped.'], as_index=False
    ).agg({'Paletes validados': 'sum', 'Valor Aprovado': 'sum'}).assign(**{'Nota fiscal': 'Total'})

    df_plt = pd.concat([df_plt, df_total], ignore_index=True)

    df_plt = df_plt.sort_values(by=['Solicitação', 'Cod.Cliente', 'Dt/ Entrega Cliente', 'Nota fiscal', 'Tp.exped.'])

    # Arredondando 'Paletes validados' para cima quando a 'Nota fiscal' for 'Total'
    df_plt.loc[df_plt['Nota fiscal'] == 'Total', 'Paletes validados'] = df_plt.loc[df_plt['Nota fiscal'] == 'Total', 'Paletes validados'].apply(math.ceil)

    # Recalculando 'Valor Aprovado' com base nos novos valores arredondados
    df_plt.loc[df_plt['Nota fiscal'] == 'Total', 'Valor Aprovado'] = df_plt.loc[df_plt['Nota fiscal'] == 'Total', 'Paletes validados'] * valor_paletizacao

    # Recalculando o Total Aprovado (somando os valores corrigidos do Total)
    total_paletes_aprovado = df_plt[df_plt['Nota fiscal'] == 'Total']['Paletes validados'].sum()
    total_valor_aprovado = df_plt[df_plt['Nota fiscal'] == 'Total']['Valor Aprovado'].sum()

    # Ajustando o Total Aprovado na tabela
    df_plt.loc[df_plt['Nota fiscal'] == 'Total Aprovado:', 'Paletes validados'] = total_paletes_aprovado
    df_plt.loc[df_plt['Nota fiscal'] == 'Total Aprovado:', 'Valor Aprovado'] = total_valor_aprovado

    # Criando uma linha de Total Aprovado, se necessário
    if 'Total Aprovado:' not in df_plt['Nota fiscal'].values:
        total_geral = pd.DataFrame({
            'Solicitação': [None],
            'Data Validação': [None],
            'Perfil Cliente': [None],
            'Cod.Cliente': [None],
            'Descrição': [None],
            'Transportador': [None],
            'Dt/ Entrega Cliente': [None],
            'Nota fiscal': ['Total Aprovado:'],
            'Paletes validados': [total_paletes_aprovado],
            'Valor Aprovado': [total_valor_aprovado],
            'Tp.exped.': [None]
        })

    df_plt = pd.concat([df_plt, total_geral], ignore_index=True)

    # Formatando a coluna 'Valor Aprovado' para exibir valores monetários corretamente
    df_plt['Valor Aprovado'] = df_plt.apply(
        lambda row: f'R$ {row["Valor Aprovado"]:,.2f}'.replace(',', '@').replace('.', ',').replace('@', '.')
        if row['Nota fiscal'] in ['Total', 'Total Aprovado:'] else '', axis=1
    )

    colunas_ordenadas = [
        'Solicitação', 'Perfil Cliente', 'Data Validação', 'Cod.Cliente', 'Descrição', 'Tp.exped.',
        'Transportador', 'Dt/ Entrega Cliente', 'Nota fiscal', 'Paletes validados',
        'Valor Aprovado'
    ]

    df_plt = df_plt[colunas_ordenadas]
# Criar DataFrame para NFs avaliadas
    df_nfs_avaliadas = pd.DataFrame(nfs_avaliadas, columns=['Nota fiscal'])

##############################
    colunas_desejadas = [
        "Localorg.transporte", "Cod.Cliente", "Descrição", "Cidade", "U.F.", "NotaFiscal", "Série",
        "Material", "Descrição2", "Quant.Fornecida", "ValorNF", "Volume", "VolumeNF", "Peso",
        "Nro.DTEmbarque", "DscTrans.Entr.Cliente", "Dt.EntregaCliente", "Tp.exped.", "Proc.especial"
    ]

###############################

    # Caminhos para os arquivos
    
    caminho_excel = r"G:"
    caminho_txt_plt = r"G:"
    caminho_txt_proc = r"G:"

    # Exportar para Excel

    with pd.ExcelWriter(caminho_excel, engine='xlsxwriter') as writer:
        df_merged.to_excel(writer, sheet_name='Processamento_ZSD223', index=False)
        df_plt.to_excel(writer, sheet_name='Calculo Palete', index=False)
        df_nfs_avaliadas.to_excel(writer, sheet_name='nf avaliadas', index=False)

    # Fechar a conexão com o banco
    conn.close()

########################## Formatação ##########################
    # Abrir o arquivo para formatação
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb["Calculo Palete"]

    # Definir a formatação
    bold_font_yellow = Font(bold=True, color="FFFF00")
    blue_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")

    linha_total_aprovado = None

    # Procurar pela linha que contém "Total Aprovado:"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if str(cell.value).strip() == "Total Aprovado:":
                linha_total_aprovado = row
                break
        if linha_total_aprovado:
            break

    if linha_total_aprovado:
        # Salvar os valores da linha "Total Aprovado"
        valores_total_aprovado = [cell.value for cell in linha_total_aprovado]

        # Remover a linha original
        ws.delete_rows(linha_total_aprovado[0].row, 1)

        # Inserir uma nova linha no topo
        ws.insert_rows(1)

        # Escrever os valores no topo com formatação condicional
        for col_index, value in enumerate(valores_total_aprovado, start=1):
            cell = ws.cell(row=1, column=col_index, value=value)
            
            # Aplicar formatação apenas ao texto "Total Aprovado:" e às células com valor (não vazias)
            if value == "Total Aprovado:" or value not in (None, "", " "):
                cell.font = bold_font_yellow
                cell.fill = blue_fill

           
    # Definir estilos
    bold_font_white = Font(bold=True, color="FFFFFF")
    bold_font_blue = Font(bold=True, color="0000FF")
    bold_font_yellow = Font(bold=True, color="FFFF00")
    blue_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Identificar a posição da coluna "Nota fiscal"
    col_nota_fiscal = None
    for col_idx, cell in enumerate(ws[2], start=1):
        if cell.value == "Nota fiscal":
            col_nota_fiscal = col_idx
            break

    if col_nota_fiscal is None:
        raise ValueError("Coluna 'Nota fiscal' não encontrada no arquivo Excel.")

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Aplicar formatação ao cabeçalho
    for cell in ws[2]:
        cell.font = bold_font_white
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    # Aplicar bordas e formatações por linha
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        nota_fiscal = str(row[col_nota_fiscal - 1].value or "").strip()

        for cell in row:
            cell.border = border_style

        if nota_fiscal == "Total":
            for cell in row:
                cell.font = bold_font_blue
                cell.fill = gray_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        
###################### Mensagem gerar arquivo ######################

    # Salvar o arquivo formatado em Excel
    print("Salvando Excel...")
    wb.save(caminho_excel)

    # Exportar para TXT
    print("Exportando para TXT...")

    # Exportar aba "Calculo Palete" para TXT
    df_plt.to_csv(caminho_txt_plt, sep="|", index=False, encoding="utf-8")

        # Exportar aba "Processamento_ZSD223" para TXT
    # df_merged.to_csv(caminho_txt_proc, sep="|", index=False, encoding="utf-8")
    df_merged[colunas_desejadas].to_csv(caminho_txt_proc, sep="|", index=False, encoding="utf-8")

    print("Processo concluído!")

if __name__ == "__main__":
    pass