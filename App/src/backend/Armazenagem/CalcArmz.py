import pandas as pd
import numpy as np
import os
from io import StringIO
from pyxlsb import open_workbook
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font
import re
from datetime import datetime
import win32api
import glob
import win32con

def main(transportadora, tp_exped):

    # === 1. Lendo Planilha Excel ===
    folder_path = r"G:"
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

    if not excel_files:
        raise FileNotFoundError("Nenhum arquivo .xlsx encontrado na pasta.")

    latest_file = max(excel_files, key=os.path.getmtime)
    df = pd.read_excel(latest_file, sheet_name=0, skiprows=6, usecols="B:J", engine="openpyxl")

    df = df.dropna(how='all')  # Remove linhas vazias
    df.columns = df.columns.str.strip().str.lower()

    # Padroniza dados
    df["nota fiscal"] = df["nota fiscal"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
    df["série"] = df["série"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
    df["nfe_série"] = df["nota fiscal"] + "-" + df["série"]

    print(df.columns)

    def converter_data_mista(valor):
            try:
                if pd.notna(valor) and isinstance(valor, (int, float)):
                    return pd.to_datetime(valor, origin='1899-12-30', unit='D')
                else:
                    return pd.to_datetime(valor, dayfirst=True, errors='coerce')
            except:
                return pd.NaT
    df["data início"] = df["data início"].apply(converter_data_mista)
    df["data fim"] = df["data fim"].apply(converter_data_mista)

    # === 2. Lendo arquivo TXT ===
    source_txt = os.path.join(folder_path, "zsd223.txt")
    data = []
    header_line = None

    try:
        with open(source_txt, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line):
                header_line = i
                break

        if header_line is not None:
            header = [col.strip().lower() for col in lines[header_line].split("|")[1:-1]]
            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line):
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    if row and row[0] != "*":
                        data.append(row)

            df_1 = pd.DataFrame(data, columns=header)
            df_1.columns = df_1.columns.str.replace(' ', '').str.strip().str.lower()

            # Padroniza dados
            df_1["notafiscal"] = df_1["notafiscal"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
            df_1["série"] = df_1["série"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
            df_1["nfe_série"] = df_1["notafiscal"] + "-" + df_1["série"]

        else:
            df_1 = pd.DataFrame()

    except Exception as e:
        print("Erro ao processar o arquivo TXT:", e)
        df_1 = pd.DataFrame()

        print(df_1_selected.columns.tolist())  # Para verificar se 'peso' e 'volume' estão realmente presentes

    # Seleciona colunas
    df_selected = df[[
        'nota fiscal', 'série', 'nome destinatário', 'cidade', 'u.f',
        'tipo da nf', 'peso bruto (kg)', 'data início', 'data fim', 'nfe_série'
    ]]

    df_selected = df_selected.rename(columns={
        'cidade': 'Cidade_1',
        'u.f': 'U.F_1', 
        'peso bruto (kg)': 'Peso Bruto (KG)_1'
    })

    df_1_selected = df_1[[
        'descrição', 'cidade', 'u.f.', 'notafiscal', 'peso', 'volume',
        'dsctrans.entr.cliente', 'dt.previstadeentrega', 'dt.entregacliente', 'nfe_série', 'tp.exped.'
    ]]

    df_1_selected = df_1_selected.rename(columns={
        'cidade': 'Cidade',
        'u.f.': 'U.F'
    })

    # Merge
    df_final = pd.merge(df_selected, df_1_selected, on='nfe_série', how='inner')

    print(df_final.columns)

    df_final = df_final.rename(columns={
        "nota fiscal": "Nota fiscal"})

###################### NF JÁ AVALIADAS ######################

    def exibir_mensagem_nfs(nf_calculadas):
        """Exibe uma caixa de mensagem com as NFs já calculadas."""
  
        mensagem = "As seguintes Notas Fiscais já foram calculadas e estão no banco de dados:\n" + "\n".join(nf_calculadas)
        win32api.MessageBox(0, mensagem, "NFs Já Calculadas", win32con.MB_ICONINFORMATION)

    # Função para limpar e normalizar as Notas Fiscais
    def limpar_nf(nf):
        """Remove espaços extras e converte para maiúsculas"""
        return str(nf).strip().upper()

    # Caminho para o banco de dados
    source = r"G:"

    def verifica_nf_calculadas(df, conn):
        """
        Verifica se as NFs do DataFrame já foram calculadas e estão armazenadas no banco de dados.
        """
        # Obter as NFs existentes no banco de dados
        query = "SELECT DISTINCT `Nota fiscal` FROM BI_Armazenagem"
        nf_existentes = pd.read_sql_query(query, conn)['Nota fiscal'].dropna().tolist()
        
        # Normalizar as NFs do banco
        nf_existentes = [limpar_nf(nf) for nf in nf_existentes]
        
        # Normalizar as NFs no DataFrame
        df['Nota fiscal'] = df['Nota fiscal'].apply(limpar_nf)
        
        # Verificar se há NFs duplicadas
        nf_calculadas = df[df['Nota fiscal'].isin(nf_existentes)]
        
        if not nf_calculadas.empty:
            # Exibir a mensagem usando a função de toast
            exibir_mensagem_nfs(nf_calculadas['Nota fiscal'].unique())
            # Remover as NFs já calculadas do DataFrame
            df = df[~df['Nota fiscal'].isin(nf_existentes)]
        
        return df, nf_calculadas['Nota fiscal'].unique()

    # Conectar ao banco de dados
    conn = sqlite3.connect(source)

    # Verificar NFs já calculadas antes de gerar o arquivo
    df_final, nfs_avaliadas = verifica_nf_calculadas(df_final, conn)

    if df_final.empty:
        print("Todas as Notas Fiscais já foram calculadas.")
    else:
        print("Continuando o processo de geração de arquivo...")

###################### Cálculos ######################

    df_final['Data Requisição'] = datetime.today().strftime('%d/%m/%Y')
    # df_final["Ref. Requisição"] = 78

    
  # REQ
        # Função para encontrar o próximo nome de arquivo disponível
    def get_next_filename(base_name, extension, folder_path):
        i = 0
        while os.path.exists(os.path.join(folder_path, f"{base_name + i}.{extension}")):
            i += 1
        return base_name + i

    # Base name, extensão e caminho da pasta
    base_name = 78
    extension = "txt"
    folder_path = r"G:"

    # Obter o próximo nome de arquivo disponível
    next_filename = get_next_filename(base_name, extension, folder_path)

    # Atualizar o valor de REQ
    df_final['Ref. Requisição'] = next_filename

    # CONSULTAS SQL
    conn = sqlite3.connect(r"G:")  
    transp_df = pd.read_sql("SELECT * FROM Transp", conn)

    df_final = df_final.rename(columns={"dsctrans.entr.cliente": "De"})
    df_final = df_final.merge(transp_df[['De', 'PARA']], on='De', how='left')
    df_final = df_final.rename(columns={"PARA": "Transportadora"})

    # Filtro por transportadora
    df_final = df_final[(df_final['Transportadora'] == transportadora)]
    
    # Datas e cálculo de dias
    df_final['data início'] = pd.to_datetime(df_final['data início'], errors='coerce')
    df_final['data fim'] = pd.to_datetime(df_final['data fim'], errors='coerce')

    dias = (df_final['data fim'] - df_final['data início']).dt.days + 1 - 15
    df_final['Total dias Solicitados'] = np.where(
        df_final['data fim'].isna(),
        0,
        np.where(
            dias <= 0,
            'REPROVADO',
            dias.astype(object)
        )
    )

    # Título do e-mail
    df_final["Título E-mail"] = df_final.apply(
        lambda row: "-" if pd.isna(row["Ref. Requisição"]) or row["Ref. Requisição"] == "" 
        else f'RCA - ARMAZENAGEM - {row["Transportadora"]} SOLIC.{row["Ref. Requisição"]}',
        axis=1
    )

    df_final = df_final.rename(columns={
        "peso": "Peso Bruto (KG)", 
        "volume": "Volume M³"
    })

    df_final['Peso Bruto (KG)'] = pd.to_numeric(df_final['Peso Bruto (KG)'].astype(str).str.replace(",", "."), errors='coerce')
    df_final['Volume M³'] = pd.to_numeric(df_final['Volume M³'].astype(str).str.replace(",", "."), errors='coerce')

    df_final['Volume M³'] = df_final['Volume M³'] * 250

    # Peso cubado como o maior valor entre peso e volume
    df_final['Peso Cubado'] = df_final[['Peso Bruto (KG)', 'Volume M³']].max(axis=1)

    # Renomeando outras colunas finais
    df_final = df_final.rename(columns={
        "série": "Série",
        "nome destinatário": "Nome Destinatário",
        "tipo da nf": "Tipo da NF", 
        "data início": "Data Início", 
        "data fim": "Data Fim",
        "descrição": "Destinatário",
        "dt.previstadeentrega": "Data Prevista",
        "dt.entregacliente": "Data de entrega", 
        "nfe_série": "NF"
    })

    df_final = df_final[(df_final['tp.exped.'] == tp_exped)]
    
    # if tp_exped.lower() in ['Z1', 'E1']:
    #     tabela_ft = pd.read_sql("SELECT * FROM tabelas_lotacao", conn)
    # elif tp_exped.lower() == 'Z2':
    #     tabela_ft = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)
    # else:
    #     raise ValueError("tp_exped inválido. Esperado 'Z1', 'Z2' ou 'E1'.")

    tp_exped = tp_exped.upper()

    if tp_exped in ['Z1', 'E1']:
        tabela_ft = pd.read_sql("SELECT * FROM tabelas_lotacao", conn)
    elif tp_exped == 'Z2':
        tabela_ft = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)
    else:
        raise ValueError("tp_exped inválido. Esperado 'Z1', 'Z2' ou 'E1'.")



    tabela_ft = tabela_ft.drop_duplicates(subset='TRANSPORTADORA', keep='first')
    df_final = df_final.merge(
        tabela_ft[['TRANSPORTADORA', 'Free Time']],
        left_on='Transportadora', right_on='TRANSPORTADORA',
        how='left'
    )
    df_final['Data Prevista'] = pd.to_datetime(df_final['Data Prevista'], errors='coerce', dayfirst=True)
    df_final['Início Armazenagem'] = df_final['Data Prevista'] + pd.to_timedelta(df_final['Free Time'], unit='D')
    df_final.drop(columns=['TRANSPORTADORA'], inplace=True)


    df_final['Data de entrega'] = pd.to_datetime(df_final['Data de entrega'], errors='coerce', dayfirst=True)
    df_final['Início Armazenagem'] = pd.to_datetime(df_final['Início Armazenagem'], errors='coerce')
    df_final['Dias de Armazenagem'] = (df_final['Data de entrega'] - df_final['Início Armazenagem']).dt.days
    df_final['Dias de Armazenagem'] = df_final['Dias de Armazenagem'].apply(lambda x: max(x, 0))
    
    colunas_datas = [
        "Data Início", "Data Fim", "Data Requisição", "Data Prevista", 
        "Data de entrega", "Início Armazenagem"
    ]
    for col in colunas_datas:
        if col in df_final.columns:
            df_final[col] = pd.to_datetime(df_final[col], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')

    # Valor
    print(f"tp_exped: {tp_exped}")

    if tp_exped in ['Z1', 'E1']: 
        tabela_ft = pd.read_sql("SELECT * FROM tabelas_lotacao", conn)
    elif tp_exped == 'Z2': 
        tabela_ft = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)
    else:
        raise ValueError("tp_exped inválido. Esperado 'Z1', 'Z2' ou 'E1'.")

    tabela_ft = tabela_ft.drop_duplicates(subset='TRANSPORTADORA', keep='first')

    df_final = df_final.merge(
        tabela_ft[['TRANSPORTADORA', 'TX ARMAZENAGEM']], 
        left_on='Transportadora', right_on='TRANSPORTADORA',
        how='left'
    )

    print(df_final[['Transportadora', 'TX ARMAZENAGEM']].head())



    # Cálculo de Valor Armaz
    df_final['Valor Armaz'] = (df_final['Peso Cubado'] / 1000) * df_final['Dias de Armazenagem'] * df_final['TX ARMAZENAGEM']
    print(df_final[['Valor Armaz']].head())


    # Reordenação das colunas finais
    ordem_final = [
        "Data Requisição", "Nota fiscal", "Série", "Nome Destinatário", "Cidade_1", "U.F_1",
        "Tipo da NF", "Peso Bruto (KG)_1", "Data Início", "Data Fim", "Total dias Solicitados",
        "Ref. Requisição", "Título E-mail", "Destinatário", "Cidade", "U.F", "NF",
        "Peso Bruto (KG)", "Volume M³", "Peso Cubado", "Transportadora","tp.exped.", "Data Prevista",
        "Início Armazenagem", "Data de entrega", "Dias de Armazenagem", "Valor Armaz"
    ]

    df_final = df_final[[col for col in ordem_final if col in df_final.columns]]
    
    def limpa_e_numeric(serie):
        s = serie.astype(str).str.strip()
        s = s.str.replace(r"[^0-9,\.-]", "", regex=True)
        s = s.str.replace(",", ".", regex=False)
        return pd.to_numeric(s, errors="coerce").fillna(0)
    
    colunas_valores = [
        "Valor Armaz"  
    ]
    for col in colunas_valores:
        if col in df_final.columns:
            df_final[col] = limpa_e_numeric(df_final[col])
    totais = {col: df_final[col].sum() for col in colunas_valores if col in df_final.columns}
    df_nfs_avaliadas = pd.DataFrame(nfs_avaliadas, columns=["Nota fiscal"])

    output_xlsx = r"G:"
    output_txt  = r"G:"

    with pd.ExcelWriter(output_xlsx, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, sheet_name="Relatório", index=False, startrow=1)
        df_nfs_avaliadas.to_excel(writer, sheet_name="nf avaliadas", index=False)
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Relatório"]  

    # estilos
    fill_header = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    font_yellow = Font(bold=True, color="FFFF00")
    font_branca = Font(color="FFFFFF")
    for col in colunas_valores:
        idx = df_final.columns.get_loc(col) + 1  
        cell = ws.cell(row=1, column=idx)
        texto = f"R$ {totais[col]:,.2f}"
      
        texto = texto.replace(",", "v").replace(".", ",").replace("v", ".")
        cell.value = texto
        cell.fill  = fill_header
        cell.font  = font_yellow
    for cell in ws[2]:
        cell.fill = fill_header
        cell.font = font_branca
       
    wb.save(output_xlsx)
    print(f"Arquivo Excel salvo em: {output_xlsx}")

    # 🔹 **Salvar como TXT**
    df_final.to_csv(output_txt, sep="|", index=False, encoding="utf-8")
    print(f"Arquivo TXT salvo em: {output_txt}")


if __name__ == "__main__":
    pass