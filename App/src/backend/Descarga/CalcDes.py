import pandas as pd
import numpy as np
import os
from io import StringIO
from pyxlsb import open_workbook
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font
import re
from datetime import datetime
import win32api
import glob
import win32con

def main(transportadora):

    # === 1. Excel ===

    folder_path = r"G:"

    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

    if not excel_files:
        raise FileNotFoundError("Nenhum arquivo .xlsx encontrado na pasta.")

    latest_file = max(excel_files, key=os.path.getmtime)
    df = pd.read_excel(latest_file, sheet_name=0, skiprows=7, usecols="B:J", engine="openpyxl")
    df = df.dropna(how='all')
    df.columns = df.columns.str.strip().str.lower()
    df["nota fiscal"] = df["nota fiscal"].astype(str).str.lstrip("0").str.strip()
    df["série"] = df["série"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
    df["nfe_série"] = df["nota fiscal"] + "-" + df["série"]
    # print(df.columns)

    # === 2. TXT ===
    source_txt = r"G:"
    data = []
    header_line = None

    try:
        with open(source_txt, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line):
                header_line = i
                break

        if header_line is not None:
            header = [col.strip().lower() for col in lines[header_line].split("|")[1:-1]]

            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line):
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    if row and row[0] != "*":
                        data.append(row)

            df_1 = pd.DataFrame(data, columns=header)
            df_1.columns = df_1.columns.str.replace(' ', '').str.strip().str.lower()

            df_1["notafiscal"] = df_1["notafiscal"].astype(str).str.lstrip("0").str.strip()
            df_1["série"] = df_1["série"].apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")
            df_1["nfe_série"] = df_1["notafiscal"] + "-" + df_1["série"]
        else:
            df_1 = pd.DataFrame()

    except Exception as e:
        print("Erro ao processar o arquivo TXT:", e)
        df_1 = pd.DataFrame()

    diff = set(df["nfe_série"].unique()) - set(df_1["nfe_série"].unique())

    # === Seleção e merge ===
    df_selected = df[[
        'nota fiscal', 'série', 'cliente destinatário', 'cidade', 'u.f',
        'tipo de serviço:', 'cobrança por:', 'quantidade:', 'valor unitário:', 'nfe_série'
    ]]

    df_1_selected = df_1[[
        'cod.cliente', 'descrição', 'cidade', 'u.f.', 'dsctrans.entr.cliente',
        'tp.exped.', 'dt.entregacliente', 'localdeexpedição','peso', 'nfe_série'
    ]]

    df_1_selected = df_1_selected.rename(columns={
        'cidade': 'cidade_txt',
        'u.f.': 'uf_txt',
        'peso': 'PESO (KG)', 
        'cod.cliente' : 'Código'
    })

    df_final = pd.merge(df_selected, df_1_selected, on='nfe_série', how='inner')

    # === Requisição e título ===
    df_final['Data de Requisição'] = datetime.today().strftime('%d/%m/%Y')
    # df_final["N° Requisição"] = 881


  # REQ
        # Função para encontrar o próximo nome de arquivo disponível
    def get_next_filename(base_name, extension, folder_path):
        i = 0
        while os.path.exists(os.path.join(folder_path, f"{base_name + i}.{extension}")):
            i += 1
        return base_name + i

    # Base name, extensão e caminho da pasta
    base_name = 881
    extension = "txt"
    folder_path = r"G:" 

    # Obter o próximo nome de arquivo disponível
    next_filename = get_next_filename(base_name, extension, folder_path)

    # Atualizar o valor de REQ
    df_final['N° Requisição'] = next_filename


    print(df_final.head())  #Depuração

    # === Banco de dados ===
    conn = sqlite3.connect(r"G:")  
   
    book_df = pd.read_sql("SELECT * FROM Book", conn)

    book_df['CÓDIGO'] = book_df['CÓDIGO'].astype(str)

    # Faz o merge usando os nomes corretos
    df_final = pd.merge(
        df_final,
        book_df[['CÓDIGO', 'DESCARGA POR CONTA DO DESTINATÁRIO?']],
        left_on='Código',
        right_on='CÓDIGO',
        how='left'
    )
    # Renomeia a coluna após o merge
    df_final = df_final.rename(columns={
        'DESCARGA POR CONTA DO DESTINATÁRIO?': 'Procede?'
    })
    df_final = df_final.drop(columns=['CÓDIGO'])
    print(df_final.head())  #Depuração
    # === Transportadora ===
    transp_df = pd.read_sql("SELECT * FROM Transp", conn)
    df_final = df_final.rename(columns={"dsctrans.entr.cliente": "De"})
    df_final = df_final.merge(transp_df[['De', 'PARA']], on='De', how='left')
    df_final = df_final.rename(columns={"PARA": "Transportadora"})
    df_final = df_final.drop(columns=["De"]) 

    df_final = df_final[(df_final['Transportadora'] == transportadora)]
    print(df_final.head())  #Depuração
    # === REF ===
 
    df_final["REF"] = (
        df_final["Transportadora"].astype(str).str.strip() +
        df_final["tp.exped."].astype(str).str.strip() +
        df_final["localdeexpedição"].astype(str).str.strip()
    )

    df_final["Título E-mail"] = df_final.apply(
        lambda row: "-" if pd.isna(row["N° Requisição"]) or row["N° Requisição"] == "" 
        else f'RCA - DESCARGA - {row["Transportadora"]} SOLIC.{row["N° Requisição"]}',
        axis=1
    )

    # === Valor solicitado ===
    df_final["quantidade:"] = pd.to_numeric(df_final["quantidade:"], errors="coerce")
    df_final["valor unitário:"] = pd.to_numeric(df_final["valor unitário:"], errors="coerce")
    df_final["Valor Solicitado"] = df_final["quantidade:"] * df_final["valor unitário:"]

# === Colunas renomeadas ===
    df_final = df_final.rename(columns={
        "nota fiscal": "Nota fiscal",
        "série": "Série",
        "cliente destinatário": "Cliente Destinatário",
        "cidade": "Cidade_1",
        "u.f": "U.F",
        "tipo de serviço:": "Tipo de serviço:",
        "cobrança por:": "Cobrança por:",
        "quantidade:": "Quantidade:",
        "valor unitário:": "Valor unitário:",
        "descrição": "Destinatário",
        "cidade_txt": "Cidade",
        "uf_txt": "UF",
        "tp.exped.": "Tipo de expedição",
        "dt.entregacliente": "Data de entrega",
        "localdeexpedição": "Local de Expedição",
        "nfe_série": "Nfe"
    })

###################### NF JÁ AVALIADAS ######################

    def exibir_mensagem_nfs(nf_calculadas):
        """Exibe uma caixa de mensagem com as NFs já calculadas."""
        mensagem = "As seguintes Notas Fiscais já foram calculadas e estão no banco de dados:\n" + "\n".join(nf_calculadas)
        win32api.MessageBox(0, mensagem, "NFs Já Calculadas", win32con.MB_ICONINFORMATION)
    # Função para limpar e normalizar as Notas Fiscais
    def limpar_nf(nf):
        """Remove espaços extras e converte para maiúsculas"""
        return str(nf).strip().upper()
    # Caminho para o banco de dados
    source = r"G:"
    def verifica_nf_calculadas(df, conn):
        """
        Verifica se as NFs do DataFrame já foram calculadas e estão armazenadas no banco de dados.
        """
        # Obter as NFs existentes no banco de dados
        query = "SELECT DISTINCT `Nota fiscal` FROM BI_Descarga"
        nf_existentes = pd.read_sql_query(query, conn)['Nota fiscal'].dropna().tolist()
        # Normalizar as NFs do banco
        nf_existentes = [limpar_nf(nf) for nf in nf_existentes]
        # Normalizar as NFs no DataFrame
        df['Nota fiscal'] = df['Nota fiscal'].apply(limpar_nf)
        # Verificar se há NFs duplicadas
        nf_calculadas = df[df['Nota fiscal'].isin(nf_existentes)]
        if not nf_calculadas.empty:
            # Exibir a mensagem usando a função de toast
            exibir_mensagem_nfs(nf_calculadas['Nota fiscal'].unique())
            # Remover as NFs já calculadas do DataFrame
            df = df[~df['Nota fiscal'].isin(nf_existentes)]
        return df, nf_calculadas['Nota fiscal'].unique()
    # Conectar ao banco de dados
    conn = sqlite3.connect(source)
    # Verificar NFs já calculadas antes de gerar o arquivo
    df_final, nfs_avaliadas = verifica_nf_calculadas(df_final, conn)
    if df_final.empty:
        print("Todas as Notas Fiscais já foram calculadas.")
    else:
        print("Continuando o processo de geração de arquivo...")



    df_final["Data de entrega"] = pd.to_datetime(df_final["Data de entrega"], format="%d.%m.%Y").dt.strftime("%d/%m/%Y")

    # === Chave ===
    df_final["Chave"] = (
        df_final["Código"].astype(str).str.zfill(5) + 
        pd.to_datetime(df_final["Data de entrega"], dayfirst=True).dt.strftime("%d%m%y")
)
    # === Agrupador ===
    df_final["Agrupador"] = (df_final["Chave"] != df_final["Chave"].shift(-1)).astype(int)
    df_final = df_final.sort_values(by="Chave").reset_index(drop=True)
    df_final["Agrupador"] = (df_final["Chave"] != df_final["Chave"].shift(-1)).astype(int)

    # === Valor cobrado ===
    def calcular_valor_cobrado(row, df_ref):
        if row['Procede?'] == 'não procede':
            return 0
        elif row['Agrupador'] == 1:
            return df_ref.loc[df_ref['Chave'] == row['Chave'], 'Valor Solicitado'].sum()
        else:
            return 0 
    df_final['Valor Cobrado por descarga'] = df_final.apply(lambda row: calcular_valor_cobrado(row, df_final), axis=1)
    df_final['Valor Cobrado por descarga'] = pd.to_numeric(df_final['Valor Cobrado por descarga'], errors='coerce').fillna(0)
    # df_final['PESO (KG)'] = pd.to_numeric(df_final['PESO (KG)'], errors='coerce').fillna(0)

   # === Valor Descarga Validado ===
    # frac_df = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)
    # lot_df = pd.read_sql("SELECT * FROM tabelas_lotacao", conn)
    # def obter_franquia(row):
    #     transportadora = row['Transportadora']
    #     tipo = row['Tipo de expedição']
    #     if tipo == 'Z1':
    #         result = lot_df.loc[lot_df['TRANSPORTADORA'] == transportadora, 'Franquia']
    #     elif tipo == 'Z2':
    #         result = frac_df.loc[frac_df['TRANSPORTADORA'] == transportadora, 'Franquia']
    #     else:
    #         return 0
    #     return result.values[0] if not result.empty else 0


    frac_df = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)
    lot_df = pd.read_sql("SELECT * FROM tabelas_lotacao", conn)

    def obter_franquia(row):
        transportadora = row['Transportadora']
        tipo = row['Tipo de expedição']
        
        if tipo in ['Z1', 'E1']:
            result = lot_df.loc[lot_df['TRANSPORTADORA'] == transportadora, 'Franquia']
        elif tipo == 'Z2':
            result = frac_df.loc[frac_df['TRANSPORTADORA'] == transportadora, 'Franquia']
        else:
            return 0
        
        return result.values[0] if not result.empty else 0
    

    df_final['Franquia'] = df_final.apply(obter_franquia, axis=1)
    df_final['Valor Cobrado por descarga'] = pd.to_numeric(df_final['Valor Cobrado por descarga'], errors='coerce').fillna(0)
    df_final['Franquia'] = pd.to_numeric(df_final['Franquia'], errors='coerce').fillna(0)
    def calcular_validado(row):
        if str(row['Procede?']).strip().upper() == 'NÃO PROCEDE':
            return 0
        elif row['Valor Cobrado por descarga'] == 0:
            return 0
        elif (row['Valor Cobrado por descarga'] - row['Franquia']) < 0:
            return 0
        else:
            return row['Valor Cobrado por descarga'] - row['Franquia']
    df_final['Valor Descarga Validado'] = df_final.apply(calcular_validado, axis=1)

    # === Valor Validado por NF ===

    df_final['PESO_FLOAT'] = (
        df_final['PESO (KG)']
        .astype(str)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
        .astype(float)
    )

    df_final['PESO_FORMATADO'] = df_final['PESO_FLOAT'].apply(
        lambda x: f"{x:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
    )

    soma_valores_validado = df_final.groupby('Chave')['Valor Descarga Validado'].transform('sum')
    soma_pesos = df_final.groupby('Chave')['PESO_FLOAT'].transform('sum')
    peso_atual = df_final['PESO_FLOAT']

    df_final['Valor Validado por NF'] = (soma_valores_validado * peso_atual) / soma_pesos
    df_final['Valor Validado por NF'] = df_final['Valor Validado por NF'].fillna(0)
    df_final['Valor Validado por NF'] = df_final['Valor Validado por NF'].apply(
        lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
    )

    # === Franquia ===
    df_final['Franquia?'] = df_final.apply(obter_franquia, axis=1)
    df_final['Franquia?'] = pd.to_numeric(df_final['Franquia'], errors='coerce').fillna(0)
    ultimas_linhas = df_final.groupby('Chave').tail(1).index
    df_final['Franquia?'] = 0  # Inicia com 0
    df_final.loc[ultimas_linhas, 'Franquia?'] = df_final.loc[ultimas_linhas, 'Franquia']
    df_final['Franquia?'] = df_final['Franquia?'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "v").replace(".", ",").replace("v", "."))
    print(df_final.head())  #Depuração
 # === Ordem Colunas ===
    ordem_final = [
        "Data de Requisição",
        "Nota fiscal",
        "Série",
        "Cliente Destinatário",
        "Cidade_1",
        "U.F",
        "Tipo de serviço:",
        "Cobrança por:",
        "Quantidade:",
        "Valor unitário:",
        "N° Requisição",
        "REF",
        "Título E-mail",
        "Código",
        "Destinatário",
        "Cidade",         # Cidade destino (da tabela TXT)
        "UF",
        "Transportadora",
        "Tipo de expedição",
        "Data de entrega",
        "Local de Expedição",
        "Nfe",
        "PESO (KG)",
        "Chave", 
        "Agrupador",
        "Valor Solicitado", 
        "Valor Cobrado por descarga", 
        "Valor Descarga Validado",
        "Valor Validado por NF", 
        "Procede?",
        "Franquia?"
    ]
    df_final = df_final[[col for col in ordem_final if col in df_final.columns]]
        
    df_nfs_avaliadas = pd.DataFrame(nfs_avaliadas, columns=['Nota Fiscal'])
    print(df_final.head())  #Depuração
 # === Conversão Totais ===
    def limpa_e_numeric(serie):
        s = serie.astype(str).str.strip()
        s = s.str.replace(r"[^0-9,\.-]", "", regex=True)
        s = s.str.replace(",", ".", regex=False)         
        return pd.to_numeric(s, errors="coerce").fillna(0)

    for col in [
        "Valor Cobrado por descarga",
        "Valor Descarga Validado",
        "Franquia?"
    ]:
        if col in df_final.columns:
            df_final[col] = limpa_e_numeric(df_final[col])
    cols_para_totalizar = [
        "Valor Cobrado por descarga",
        "Valor Descarga Validado",
        "Franquia?"
    ]
    totais = { col: df_final[col].sum() for col in cols_para_totalizar }

    print(df_final.head())  #Depuração

    # Caminho do arquivo Excel
    output_xlsx = r"G:"
    output_txt = r"G:"
    with pd.ExcelWriter(output_xlsx, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, sheet_name="Relatório", index=False, startrow=1)
        df_nfs_avaliadas.to_excel(writer, sheet_name="nf avaliadas", index=False)
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Relatório"]  

    # estilos
    fill_header = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    font_branca = Font(color="FFFFFF")
    bold_font_yellow = Font(bold=True, color="FFFF00")
    for col in cols_para_totalizar:
        idx = df_final.columns.get_loc(col) + 1  
        cell = ws.cell(row=1, column=idx)
        texto = f"R$ {totais[col]:,.2f}"
        # adapta para PT‑BR
        texto = texto.replace(",", "v").replace(".", ",").replace("v", ".")
        cell.value = texto
        cell.fill  = fill_header
        cell.font  = bold_font_yellow
    for cell in ws[2]:
        cell.fill = fill_header
        cell.font = font_branca

    wb.save(output_xlsx)
    print(f"Arquivo Excel salvo em: {output_xlsx}")

    # 🔹 **Salvar como TXT**
    df_final.to_csv(output_txt, sep="|", index=False, encoding="utf-8")
    print(f"Arquivo TXT salvo em: {output_txt}")


if __name__ == "__main__":
    pass