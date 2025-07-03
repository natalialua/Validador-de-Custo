import pandas as pd
import numpy as np
import os
from io import StringIO
from pyxlsb import open_workbook
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font
import re
from datetime import datetime
import win32api
import glob
import win32con

def main(transportadora, file_name):

    folder_path = r"G:"

    # 1º DataFrame - Arquivo XLSX
    # Buscar todos os arquivos .xlsx na pasta
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    if not excel_files:
        raise FileNotFoundError("Nenhum arquivo .xlsx encontrado na pasta.")
    latest_file = max(excel_files, key=os.path.getmtime)

    # Lendo o arquivo Excel
    df = pd.read_excel(latest_file, sheet_name=0, skiprows=14, usecols="B:J", engine="openpyxl")
    df = df.dropna(how='all')
    df.columns = df.columns.str.strip().str.lower()

    # Criando a coluna de chave de junção
    df["nota fiscal"] = df["nota fiscal"].astype(str).str.lstrip("0").str.strip()
    df["série"] = df["série"].astype(str).str.lstrip("0").str.strip()
    df["nfe_série"] = df["nota fiscal"] + "-" + df["série"]

    print(df.head()) #Depuração

    # 2º DataFrame - Arquivo TXT
    # Caminho da pasta
    folder_path = r"G:"

    # Buscar todos os arquivos .txt na pasta
    txt_files = glob.glob(os.path.join(folder_path, "*.txt"))
    if not txt_files:
        raise FileNotFoundError("Nenhum arquivo .txt encontrado na pasta.")
    latest_txt = max(txt_files, key=os.path.getmtime)
    source_txt = latest_txt

    # Lendo o arquivo
    data = []
    header_line = None

    try:
        with open(source_txt, "r", encoding="ISO-8859-1") as file:
            lines = file.readlines()
        if not lines:
            raise ValueError("O arquivo está vazio!")
        for i, line in enumerate(lines):
            if re.match(r"^\|.*\|$", line):
                header_line = i
                break
        if header_line is not None:
            header = [col.strip().lower() for col in lines[header_line].split("|")[1:-1]]
            for line in lines[header_line + 1:]:
                if re.match(r"^\|.*\|$", line):
                    row = [col.strip() for col in line.split("|")[1:-1]]
                    if not row[0] == "*":
                        data.append(row)
            df_1 = pd.DataFrame(data, columns=header)
            df_1.columns = df_1.columns.str.replace(' ', '').str.strip().str.lower()
            
            print(df_1.head())  #Depuração

            df_1["notafiscal"] = df_1["notafiscal"].astype(str).str.lstrip("0").str.strip()
            df_1["série"] = df_1["série"].astype(str).str.lstrip("0").str.strip()
            df_1["nfe_série"] = df_1["notafiscal"] + "-" + df_1["série"]
        else:
            df_1 = pd.DataFrame()
    except Exception as e:
        print("Erro ao processar o arquivo:", e)
        df_1 = pd.DataFrame()

    # Padronizando e renomeando colunas
    df_1 = df_1.rename(columns={"u.f.": "u.f", "tp.exped.": "Tipo Exped.", "descrição": "Destinatário", "volume":"Volume m³"})

    df["nfe_série"] = df["nfe_série"].str.strip()
    df_1["nfe_série"] = df_1["nfe_série"].str.strip()

    # Merge dos DataFrames
    df_final = df.merge(df_1, on="nfe_série", how="inner")
    df_1 = df_1.rename(columns={"série": "série_df_1"})

    print(df_final.head())  #Depuração

    # Realizar o merge novamente
    df_final = df.merge(df_1, on="nfe_série", how="inner")

    print(df_final.head())  #Depuração

    df_final = df_final.rename(columns={
        "nota fiscal": "Nota Fiscal",
        "série": "Série",
        "cliente destinatário": "Cliente Destinatário",  
        "cidade_x": "Cidade_1", 
        "cidade_y": "Cidade", 
        "u.f_x": "U.F",
        "u.f_y": "UF",
        "tipo do custo": "Tipo do Custo",
        "tipo de veículo": "Tipo de Veículo",
        "qtde veículos": "Qtde Veículos",
        "senha": "Senha",
        "cod.cliente": "Código Cliente",
        "itinerário": "Itinerário",
        "nfe_série": "NF + Série",
        "dt.entregacliente": "Data Entrega",
    })

###################### NF JÁ AVALIADAS ######################

    def exibir_mensagem_nfs(nf_calculadas):
        """Exibe uma caixa de mensagem com as NFs já calculadas."""
  
        mensagem = "As seguintes Notas Fiscais já foram calculadas e estão no banco de dados:\n" + "\n".join(nf_calculadas)
        win32api.MessageBox(0, mensagem, "NFs Já Calculadas", win32con.MB_ICONINFORMATION)

    # Função para limpar e normalizar as Notas Fiscais
    def limpar_nf(nf):
        """Remove espaços extras e converte para maiúsculas"""
        return str(nf).strip().upper()

    # Caminho para o banco de dados
    source = r"G:"

    def verifica_nf_calculadas(df, conn):
        """
        Verifica se as NFs do DataFrame já foram calculadas e estão armazenadas no banco de dados.
        """
        # Obter as NFs existentes no banco de dados
        query = "SELECT DISTINCT `Nota Fiscal` FROM BI_Dedicado"
        nf_existentes = pd.read_sql_query(query, conn)['Nota Fiscal'].dropna().tolist()
        
        # Normalizar as NFs do banco
        nf_existentes = [limpar_nf(nf) for nf in nf_existentes]
        
        # Normalizar as NFs no DataFrame
        df['Nota Fiscal'] = df['Nota Fiscal'].apply(limpar_nf)
        
        # Verificar se há NFs duplicadas
        nf_calculadas = df[df['Nota Fiscal'].isin(nf_existentes)]
        
        if not nf_calculadas.empty:
            # Exibir a mensagem usando a função de toast
            exibir_mensagem_nfs(nf_calculadas['Nota Fiscal'].unique())
            # Remover as NFs já calculadas do DataFrame
            df = df[~df['Nota Fiscal'].isin(nf_existentes)]
        
        return df, nf_calculadas['Nota Fiscal'].unique()

    # Conectar ao banco de dados
    conn = sqlite3.connect(source)

    # Verificar NFs já calculadas antes de gerar o arquivo
    df_final, nfs_avaliadas = verifica_nf_calculadas(df_final, conn)

    if df_final.empty:
        print("Todas as Notas Fiscais já foram calculadas.")
    else:
        print("Continuando o processo de geração de arquivo...")

    print(df_final.head())  #Depuração

    ####CALCULOS#####

    df_final['Data Entrega'] = pd.to_datetime(df_final['Data Entrega'], errors='coerce', dayfirst=True)
    df_final['Data Entrega'] = df_final['Data Entrega'].fillna(pd.Timestamp('1900-01-01'))
    df_final['Data Entrega'] = pd.to_datetime(df_final['Data Entrega'], errors='coerce', dayfirst=True)
    df_final['dt.numero'] = (df_final['Data Entrega'] - pd.Timestamp('1900-01-01')).dt.days + 2

    # Criando a Chave de Entrega 
    df_final["Chave de Entrega"] = df_final["dt.numero"].astype(str) + df_final["Código Cliente"].astype(str)
    # Formatando a 'Data Entrega'
    df_final['Data Entrega'] = df_final['Data Entrega'].dt.strftime('%d/%m/%Y')
    df_final = df_final.drop(columns=['dt.numero'])

    # REQ
        # Função para encontrar o próximo nome de arquivo disponível
    def get_next_filename(base_name, extension, folder_path):
        i = 0
        while os.path.exists(os.path.join(folder_path, f"{base_name + i}.{extension}")):
            i += 1
        return base_name + i

    # Base name, extensão e caminho da pasta
    base_name = 290
    extension = "txt"
    folder_path = r"G:"  
    # Obter o próximo nome de arquivo disponível
    next_filename = get_next_filename(base_name, extension, folder_path)
    # Atualizar o valor de REQ
    df_final['Requisição'] = next_filename

    print(df_final.head())  #Depuração

    df_final["Qtd Volume"] = (df_final["Chave de Entrega"] != df_final["Chave de Entrega"].shift(-1)).astype(int)
    df_final["Volume m³"] = df_final["Volume m³"].str.replace(',', '.').astype(float)
    df_final["Soma Vol m³"] = df_final.groupby("Chave de Entrega")["Volume m³"].cumsum()
    df_final["Soma Vol m³"] = df_final["Soma Vol m³"] * (df_final["Qtd Volume"] > 0)
    df_final["Soma Vol m³"] = df_final.groupby("Chave de Entrega")["Soma Vol m³"].transform("max")
    df_final.loc[df_final["Qtd Volume"] == 0, "Soma Vol m³"] = 0

    df_final['Data Requisição'] = datetime.today().strftime('%d/%m/%Y')
    
    # Conectar ao banco SQLite
    conn = sqlite3.connect(r"G:")  
    cursor = conn.cursor()

    # Verificando as colunas da tabela 'valor_dedicado'
    cursor.execute("PRAGMA table_info('valor_dedicado');")
    colunas = cursor.fetchall()

    # Obtendo os dados da tabela 'valor_dedicado'
    cursor.execute("SELECT * FROM valor_dedicado")
    valor_dedicado_data = cursor.fetchall()

    # Convertendo os dados para DataFrame
    valor_dedicado_df = pd.DataFrame(valor_dedicado_data, columns=["Tipo veiculo", "min", "max", "valor"])

    # Convertendo as colunas 'min' e 'max' para tipo float para evitar problemas de tipo
    valor_dedicado_df["min"] = pd.to_numeric(valor_dedicado_df["min"], errors="coerce")
    valor_dedicado_df["max"] = pd.to_numeric(valor_dedicado_df["max"], errors="coerce")

    print("Dados do valor_dedicado_df:")
    print(valor_dedicado_df)

    print(df_final.head())   #Depuração

    def validar_tipo_veiculo(volume):
        volume = pd.to_numeric(str(volume).replace(',', '.'), errors='coerce')
        if pd.isna(volume): 
            return None

        print(f"Validando volume: {volume}")

        # Se for menor que o mínimo esperado (1), assume-se "FIORINO"
        if volume < 1:
            print(f"Volume {volume} é menor que 1, assumindo 'FIORINO'")
            return "FIORINO"
        # Se for maior que o máximo esperado (55), assume-se "CARRETA"
        if volume > 55:
            print(f"Volume {volume} é maior que 55, assumindo 'CARRETA'")
            return "CARRETA"

        # Realizar a comparação de volume com os intervalos de min e max
        for _, row in valor_dedicado_df.iterrows():
            if pd.notna(row["min"]) and pd.notna(row["max"]):  
                if row["min"] <= volume <= row["max"]:
                    print(f"Volume {volume} está entre {row['min']} e {row['max']}, Tipo veiculo: {row['Tipo veiculo']}")
                    return row["Tipo veiculo"]

        return None  
    
    print(df_final.head())  #Depuração

    df_final["Tipo de veículo validado"] = df_final["Soma Vol m³"].apply(validar_tipo_veiculo)
    print(df_final.head())

    print("Dados do df_final após validação:")
    print(df_final[["Soma Vol m³", "Tipo de veículo validado"]])

    # Obtendo os dados da tabela 'Transp'
    transp_df = pd.read_sql("SELECT * FROM Transp", conn)
    print(transp_df["De"].value_counts())

    # Realizando o merge entre df_final e transp_df com base na coluna 'De'
    df_final = df_final.rename(columns={"dsctrans.entr.cliente": "De"})
    df_final = df_final.merge(transp_df[['De', 'PARA']], on='De', how='left')

    # Renomeando a coluna 'PARA' para 'Transportador'
    df_final = df_final.rename(columns={"PARA": "Transportador"})

    print(df_final.head())  #Depuração

    df_final = df_final[(df_final['Transportador'] == transportadora)]

    # Reorganizando as colunas
    ordem_final = [
        "Data Requisição", "Nota Fiscal", "Série", "Cliente Destinatário", "Cidade_1", "U.F",
        "Tipo do Custo", "Tipo de Veículo", "Qtde Veículos", "Senha","Requisição", 
        "Código Cliente", "Destinatário", "Cidade", "UF", "Itinerário", "NF + Série", 
        "Transportador", "Data Entrega", "Tipo Exped.","Chave de Entrega", "Volume m³", 
        "Qtd Volume","Soma Vol m³", "Tipo de veículo validado", "Valor Validado "
    ]

    df_final = df_final[[col for col in ordem_final if col in df_final.columns]]

    print(df_final.columns)
    print(df_final.head())  #Depuração

    # Obtendo os dados da tabela 'tabelas_fracionado'
    tabelas_fracionado_df = pd.read_sql("SELECT * FROM tabelas_fracionado", conn)

    # 🔹 Mapeamento de veículos
    mapeamento_veiculos = {
        "FIORINO": "DEDICADO FIORINO",
        "VAN": "DEDICADO VAN",
        "VUC": "DEDICADO VUC",
        "3/4": "DEDICADO 3/4",
        "TOCO": "DEDICADO TOCO",
        "TRUCK": "DEDICADO TRK",
        "CARRETA": "DEDICADO CAR"
    }

    df_final["Valor Validado "] = None

    for chave in df_final["Chave de Entrega"].unique():
        subset = df_final[df_final["Chave de Entrega"] == chave]
        
        idx = subset.index[subset["Qtd Volume"] > 0].min()
        
        if not pd.isna(idx): 
            transportadora = df_final.loc[idx, "Transportador"]
            tipo_veiculo = df_final.loc[idx, "Tipo de veículo validado"]
            coluna_veiculo = mapeamento_veiculos.get(tipo_veiculo)
            
            if transportadora in tabelas_fracionado_df["TRANSPORTADORA"].values and coluna_veiculo:
                valor = tabelas_fracionado_df.loc[tabelas_fracionado_df["TRANSPORTADORA"] == transportadora, coluna_veiculo].values[0]
                df_final.at[idx, "Valor Validado"] = valor  


    df_nfs_avaliadas = pd.DataFrame(nfs_avaliadas, columns=['Nota Fiscal'])

    print(df_final.head())  #Depuração

    conn.close()


        
    def limpa_e_numeric(serie):
            s = serie.astype(str).str.strip()
            s = s.str.replace(r"[^0-9,\.-]", "", regex=True)
            s = s.str.replace(",", ".", regex=False)
            return pd.to_numeric(s, errors="coerce").fillna(0)
        
    colunas_valores = [
        "Valor Validado "  
    ]

    # 🔹 Converte as colunas numéricas
    for col in colunas_valores:
        if col in df_final.columns:
            df_final[col] = limpa_e_numeric(df_final[col])

    # 🔹 Calcula totais das colunas válidas
    totais = {col: df_final[col].sum() for col in colunas_valores if col in df_final.columns}


    # Caminho do arquivo Excel
    output_xlsx = r"G:"
    output_txt = r"G:"

    # 🔹 **Salvar `df_final` na aba principal**
    with pd.ExcelWriter(output_xlsx, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, sheet_name="Relatório", startrow=1, index=False)
        df_nfs_avaliadas.to_excel(writer, sheet_name="nf avaliadas", index=False)

    # 🔹 **Aplicar formatação 
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Relatório"]  

    fill_header = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    font_branca = Font(color="FFFFFF")
    font_yellow = Font(bold=True, color="FFFF00")
    for col in colunas_valores:
        idx = df_final.columns.get_loc(col) + 1  
        cell = ws.cell(row=1, column=idx)
        texto = f"R$ {totais[col]:,.2f}"
        # adapta para PT‑BR
        texto = texto.replace(",", "v").replace(".", ",").replace("v", ".")
        cell.value = texto
        cell.fill  = fill_header
        cell.font  = font_yellow
    for cell in ws[2]:
        cell.fill = fill_header
        cell.font = font_branca  

    wb.save(output_xlsx)
    print(f"Arquivo Excel salvo em: {output_xlsx}")

    # 🔹 **Salvar como TXT**
    df_final.to_csv(output_txt, sep="|", index=False, encoding="utf-8")
    print(f"Arquivo TXT salvo em: {output_txt}")

if __name__ == "__main__":
    pass


